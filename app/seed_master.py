"""마스터 데이터 전체 적재 — 연관관계 포함.

실행:
  python app/seed_master.py
  또는
  python -m app.seed_master
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).parent.parent))
from app.db import DATABASE_URL

# ─────────────────────────────────────────────────────────────────────────────
# DDL: 신규 룩업 테이블
# ─────────────────────────────────────────────────────────────────────────────

NEW_TABLES_DDL = [
    """
    CREATE TABLE IF NOT EXISTS mst_neuro_level (
        nerve_root  VARCHAR(8)   PRIMARY KEY,
        myotome     VARCHAR(128) NOT NULL,
        dermatome   VARCHAR(128),
        reflex      VARCHAR(64),
        region      VARCHAR(16)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_mmt_grade (
        grade       SMALLINT     PRIMARY KEY,
        label       VARCHAR(16)  NOT NULL,
        description VARCHAR(128)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_rom_reference (
        id         SERIAL      PRIMARY KEY,
        joint      VARCHAR(32) NOT NULL,
        motion     VARCHAR(64) NOT NULL,
        normal_rom VARCHAR(32) NOT NULL,
        UNIQUE (joint, motion)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_subjective_vocab (
        id       SERIAL      PRIMARY KEY,
        axis     VARCHAR(32) NOT NULL,
        term_ko  VARCHAR(64) NOT NULL,
        term_en  VARCHAR(64),
        note     VARCHAR(128),
        UNIQUE (axis, term_ko)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_red_flag (
        id       SERIAL     PRIMARY KEY,
        category VARCHAR(32) NOT NULL,
        sign     TEXT        NOT NULL,
        action   TEXT,
        UNIQUE (category, sign)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_imaging_finding (
        id       SERIAL     PRIMARY KEY,
        modality VARCHAR(8) NOT NULL CHECK (modality IN ('X-ray','MRI','US','CT')),
        finding  VARCHAR(128) NOT NULL,
        term_en  VARCHAR(128),
        UNIQUE (modality, finding)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_grading_scale (
        scale  VARCHAR(64) PRIMARY KEY,
        target VARCHAR(64),
        grades VARCHAR(32),
        note   VARCHAR(128),
        src    VARCHAR(8)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_outcome_scale (
        scale         VARCHAR(64) PRIMARY KEY,
        target_region VARCHAR(64),
        score_range   VARCHAR(32),
        note          VARCHAR(128)
    )
    """,
    """
    CREATE TABLE IF NOT EXISTS mst_peripheral_nerve (
        id         SERIAL       PRIMARY KEY,
        nerve      VARCHAR(64)  NOT NULL,
        site       VARCHAR(128),
        symptom    VARCHAR(255),
        exam       VARCHAR(128),
        related_dx VARCHAR(128)
    )
    """,
]

# ─────────────────────────────────────────────────────────────────────────────
# 진단명 → KCD 코드 매핑 (rel_diagnosis_exam 용)
# ─────────────────────────────────────────────────────────────────────────────

DX_NAME_TO_CODE: dict[str, str] = {
    "요추 신경근병증(HNP)":       "M511",
    "척추관 협착증":               "M4806",
    "경추 신경근병증":             "M501",
    "동결견(오십견)":              "M750",
    "회전근개 파열":               "M751",
    "어깨 충돌증후군":             "M754",
    "어깨충돌증후군":              "M754",
    "외측상과염":                  "M771",
    "수근관증후군":                "G560",
    "드꿰르뱅 건초염":             "M654",
    "반월상연골 파열":             "S8329",
    "전방십자인대 파열":           "S8352",
    "내측측부인대 손상":           "S8343",
    "발목 외측인대 손상":          "S9348",
    "족저근막염":                  "M722",
    "이상근증후군":                "M79158",
    "천장관절통":                  "S336",
    "천장관절 기능장애":           "S336",
    "고관절 관절내병변":           "M1995",
    "SLAP 병변":                   "S434",
    "어깨 불안정성(전방)":         "S4309",
    "팔꿈치터널증후군":            "G562",
    "내측상과염":                  "M770",
    "척추분리증/전방전위증":       "M4316",
    "대퇴비구충돌증후군(FAI)":     "M1995",
    "아킬레스건염":                "M766",
    "중둔근 약화/GTPS":            "M707",
}

# exam 이름 정규화 alias (Excel 표기 → DB name 매핑)
EXAM_ALIASES: dict[str, str] = {
    "SLR test":                "SLR test",
    "Neurogenic claudication": "Claudication",
    "Spurling test":           "Spurling test",
    "ROM(외회전)":             "ROM",
    "Empty can":               "Empty can test",
    "Drop arm":                "Drop arm test",
    "Neer":                    "Neer test",
    "Neer sign":               "Neer sign",
    "Hawkins":                 "Hawkins test",
    "Hawkins-Kennedy test":    "Hawkins-Kennedy test",
    "Cozen":                   "Cozen test",
    "Mill":                    "Mill test",
    "Phalen":                  "Phalen test",
    "Tinel":                   "Tinel sign (wrist)",
    "Tinel(elbow)":            "Tinel (elbow)",
    "Durkan":                  "Durkan test",
    "Finkelstein":             "Finkelstein test",
    "McMurray":                "McMurray test",
    "Apley":                   "Apley grind test",
    "Lachman":                 "Lachman test",
    "Anterior drawer":         "Anterior drawer",
    "Valgus stress":           "Valgus stress test",
    "Talar tilt":              "Talar tilt test",
    "Ankle anterior drawer":   "Ankle anterior drawer",
    "Windlass test":           "Windlass test",
    "Piriformis":              "Piriformis test",
    "FAIR":                    "Piriformis test",
    "FADIR":                   "FADIR test",
    "Log roll":                "Log roll test",
    "O'Brien test":            "O'Brien test",
    "Apprehension test":       "Apprehension test",
    "Relocation test":         "Relocation test",
    "Elbow flexion test":      "Elbow flexion test",
    "Stork test(one-leg hyperextension)": "Stork test",
    "Gaenslen test":           "Gaenslen test",
    "ASIS Compression":        "SI compression",
    "Thigh thrust":            "Thigh thrust",
    "Scour test":              "Scour test",
    "Trendelenburg test":      "Trendelenburg test",
    "SI provocation cluster":  None,  # 복합 클러스터, 개별 검사로 분해 불가
    "아킬레스건 압통 촉진":    None,
    "Royal London Hospital test": None,
    "내측상과 압통":           None,
    "저항 손목 굴곡 통증":     None,
}

# ─────────────────────────────────────────────────────────────────────────────
# rel_part_diagnosis — 부위 노드 ↔ 진단 (지식 기반)
# (body_part.name_ko 정확 일치, diagnosis.code)
# ─────────────────────────────────────────────────────────────────────────────

PART_DIAGNOSIS: list[tuple[str, str]] = [
    # 무릎 구조물
    ("전방십자인대(ACL)",             "S8352"),
    ("후방십자인대(PCL)",             "S8353"),
    ("내측측부인대(MCL)",             "S8343"),
    ("내측반월상연골",                "S8329"),
    ("외측반월상연골",                "S8329"),
    ("무릎",                          "M170"),
    ("무릎",                          "M171"),
    ("무릎",                          "M179"),
    ("내측 관절선",                   "S8343"),
    ("외측 관절선",                   "S8329"),
    # 어깨
    ("회전근개",                      "M751"),
    ("회전근개",                      "S4608"),
    ("관절와순",                      "S4309"),
    ("관절와순",                      "S434"),
    ("견봉하낭(subacromial bursa)",   "M754"),
    ("상완이두건 장두",               "M752"),
    ("견봉쇄골관절(AC joint)",        "S435"),
    ("어깨/견갑대",                   "M750"),
    ("이두근고랑",                    "M752"),
    # 경추
    ("경추",                          "M501"),
    ("경추",                          "M4782"),
    ("경추",                          "M4802"),
    # 요추
    ("요추",                          "M511"),
    ("요추",                          "M4806"),
    ("요추",                          "M4316"),
    ("요추",                          "M5456"),
    # 팔꿈치
    ("외측상과",                      "M771"),
    ("내측상과",                      "M770"),
    ("주관절 척골신경구(cubital)",    "G562"),
    # 손목/손
    ("수근관",                        "G560"),
    ("요골 경상돌기",                 "M654"),
    ("TFCC 부위",                     "M2423"),
    # 발목/발
    ("전거비인대(ATFL)",              "S9348"),
    ("족저근막",                      "M722"),
    ("아킬레스건 중간부",             "M766"),
    ("아킬레스건 부착부",             "M766"),
    ("종골 내측결절",                 "M722"),
    # 고관절/골반
    ("천장관절(SI joint)",            "S336"),
    ("이상근",                        "M79158"),
    ("서혜부",                        "M1995"),
    ("전자낭(trochanteric bursa)",    "M707"),
    ("중둔근(gluteus medius)",        "M79158"),
]

# ─────────────────────────────────────────────────────────────────────────────
# 헬퍼
# ─────────────────────────────────────────────────────────────────────────────

def iter_rows(ws, *, skip_header: bool = True):
    it = ws.iter_rows(values_only=True)
    if skip_header:
        next(it)
    return list(it)


def _resolve_exam(token: str, name_to_id: dict[str, int]) -> int | None:
    token = token.strip()
    if not token:
        return None
    # 1) 명시적 alias
    canonical = EXAM_ALIASES.get(token)
    if canonical is None and token in EXAM_ALIASES:
        return None  # 의도적 skip
    if canonical:
        return name_to_id.get(canonical)
    # 2) exact
    if token in name_to_id:
        return name_to_id[token]
    # 3) case-insensitive
    tl = token.lower()
    for name, eid in name_to_id.items():
        if name.lower() == tl:
            return eid
    # 4) 공통 단어 매칭 (4자 이상 단어 기준)
    t_words = {w for w in tl.split() if len(w) >= 4}
    if not t_words:
        return None
    for name, eid in name_to_id.items():
        n_words = {w for w in name.lower().split() if len(w) >= 4}
        if t_words & n_words:
            return eid
    return None


def _split_exam_tokens(raw: str) -> list[str]:
    """'Lachman / Anterior drawer, SLR test' → ['Lachman', 'Anterior drawer', 'SLR test']"""
    tokens = []
    for part in raw.replace(" / ", ",").split(","):
        t = part.strip()
        if t:
            tokens.append(t)
    return tokens


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = openpyxl.load_workbook("master_data/orthopedic_master_data.xlsx")
    engine = create_engine(DATABASE_URL)

    with engine.begin() as conn:

        # ── 1. 신규 룩업 테이블 DDL ──────────────────────────────────────────
        print("▶ 1. 신규 룩업 테이블 DDL 적용…")
        for ddl in NEW_TABLES_DDL:
            conn.execute(text(ddl))

        # ── 2. IDENTITY → BY DEFAULT ──────────────────────────────────────────
        print("▶ 2. IDENTITY BY DEFAULT 설정…")
        conn.execute(text(
            "ALTER TABLE mst_body_part ALTER COLUMN id SET GENERATED BY DEFAULT"
        ))
        conn.execute(text(
            "ALTER TABLE mst_exam_item ALTER COLUMN item_id SET GENERATED BY DEFAULT"
        ))

        # ── 3. 기존 마스터 테이블 초기화 (역FK 순서) ────────────────────────
        print("▶ 3. 마스터 테이블 초기화…")
        CLEAR_ORDER = [
            "rel_structure_exam", "rel_part_diagnosis",
            "rel_diagnosis_exam", "rel_diagnosis_order",
            "map_alias", "mst_exam_item", "mst_body_part",
            "mst_value_scale", "laterality",
            "mst_neuro_level", "mst_mmt_grade", "mst_rom_reference",
            "mst_subjective_vocab", "mst_red_flag", "mst_imaging_finding",
            "mst_grading_scale", "mst_outcome_scale", "mst_peripheral_nerve",
        ]
        for t in CLEAR_ORDER:
            conn.execute(text(f"DELETE FROM {t}"))

        # ── 4. laterality ─────────────────────────────────────────────────────
        print("▶ 4. laterality 삽입…")
        ws = wb["laterality"]
        lat_n = 0
        for r in iter_rows(ws):
            code, label_ko, *_ = r
            if code:
                conn.execute(text(
                    "INSERT INTO laterality(code, label) "
                    "VALUES(:c, :l) ON CONFLICT DO NOTHING"
                ), {"c": code, "l": label_ko})
                lat_n += 1
        print(f"   laterality: {lat_n}건")

        # ── 5. mst_value_scale ────────────────────────────────────────────────
        print("▶ 5. mst_value_scale 삽입…")
        ws = wb["value_scale"]
        vs_n = 0
        for r in iter_rows(ws):
            vtype, pattern, example, *_ = r
            if vtype:
                conn.execute(text(
                    "INSERT INTO mst_value_scale(value_type, pattern, example) "
                    "VALUES(:v, :p, :e) ON CONFLICT DO NOTHING"
                ), {"v": vtype, "p": pattern, "e": example})
                vs_n += 1
        print(f"   mst_value_scale: {vs_n}건")

        # ── 6. mst_body_part (위상 정렬 후 삽입) ─────────────────────────────
        print("▶ 6. mst_body_part 삽입…")
        ws = wb["body_part"]
        bp_all: list[dict] = []
        for r in iter_rows(ws):
            bid, parent, name_ko, name_en, node_type, *_ = r
            if bid:
                bp_all.append({
                    "id":       int(bid),
                    "parent_id": int(parent) if parent else None,
                    "name":     str(name_ko).strip(),
                    "name_en":  str(name_en).strip() if name_en else None,
                    "node_type": node_type,
                })

        # 위상 정렬
        inserted_ids: set[int] = set()
        bp_ordered: list[dict] = []
        remaining = list(bp_all)
        while remaining:
            batch = [r for r in remaining
                     if r["parent_id"] is None or r["parent_id"] in inserted_ids]
            if not batch:
                print(f"   ⚠ 순환참조 의심, 잔여 {len(remaining)}건 강제 추가")
                batch = remaining
            bp_ordered.extend(batch)
            inserted_ids.update(r["id"] for r in batch)
            inserted_set = {r["id"] for r in batch}
            remaining = [r for r in remaining if r["id"] not in inserted_set]

        for r in bp_ordered:
            conn.execute(text(
                "INSERT INTO mst_body_part(id, parent_id, name, name_en, node_type) "
                "OVERRIDING SYSTEM VALUE "
                "VALUES(:id, :parent_id, :name, :name_en, :node_type)"
            ), r)

        conn.execute(text(
            "SELECT setval(pg_get_serial_sequence('mst_body_part','id'), "
            "(SELECT MAX(id) FROM mst_body_part))"
        ))

        # 이름 → ID 역인덱스
        name_to_bp_id: dict[str, int] = {r["name"]: r["id"] for r in bp_ordered}
        print(f"   mst_body_part: {len(bp_ordered)}건")

        # ── 7. mst_exam_item ──────────────────────────────────────────────────
        print("▶ 7. mst_exam_item 삽입…")
        ws = wb["exam_item"]
        seen_exam_names: set[str] = set()
        exam_n = 0
        for r in iter_rows(ws):
            cat, name, _full, tpart, tnode, vtype, *_ = r
            if not name:
                continue
            name = str(name).strip()
            if name in seen_exam_names:   # Excel 내 중복 제거
                continue
            seen_exam_names.add(name)

            tpart_id = name_to_bp_id.get(str(tpart).strip()) if tpart else None
            tnode_id = name_to_bp_id.get(str(tnode).strip()) if tnode else None
            conn.execute(text(
                "INSERT INTO mst_exam_item(category, name, target_part, target_node, value_type) "
                "VALUES(:cat, :nm, :tp, :tn, :vt)"
            ), {
                "cat": cat, "nm": name,
                "tp": tpart_id, "tn": tnode_id,
                "vt": str(vtype).strip() if vtype else None,
            })
            exam_n += 1

        conn.execute(text(
            "SELECT setval(pg_get_serial_sequence('mst_exam_item','item_id'), "
            "(SELECT MAX(item_id) FROM mst_exam_item))"
        ))

        # item 이름 → item_id 역인덱스
        exam_name_to_id: dict[str, int] = {}
        for row in conn.execute(text("SELECT item_id, name FROM mst_exam_item")):
            exam_name_to_id[row[1].strip()] = row[0]
        print(f"   mst_exam_item: {exam_n}건")

        # ── 8. 신규 룩업 테이블 삽입 ─────────────────────────────────────────
        print("▶ 8. 신규 룩업 테이블 삽입…")

        # neuro_level
        ws = wb["neuro_level"]
        nl_n = 0
        for r in iter_rows(ws):
            nerve, myotome, dermatome, reflex, region, *_ = r
            if nerve:
                conn.execute(text(
                    "INSERT INTO mst_neuro_level(nerve_root,myotome,dermatome,reflex,region) "
                    "VALUES(:nr,:my,:dm,:rf,:rg) ON CONFLICT DO NOTHING"
                ), {"nr": nerve, "my": myotome, "dm": dermatome,
                    "rf": reflex, "rg": region})
                nl_n += 1
        print(f"   mst_neuro_level: {nl_n}건")

        # mmt_grade  (grade 열 형식: "MMT 0" 또는 정수)
        ws = wb["mmt_grade"]
        mmt_n = 0
        for r in iter_rows(ws):
            grade, label, desc, *_ = r
            if grade is not None:
                grade_str = str(grade).strip()
                grade_int = int(grade_str.split()[-1])  # "MMT 5" → 5
                conn.execute(text(
                    "INSERT INTO mst_mmt_grade(grade,label,description) "
                    "VALUES(:g,:l,:d) ON CONFLICT DO NOTHING"
                ), {"g": grade_int, "l": label, "d": desc})
                mmt_n += 1
        print(f"   mst_mmt_grade: {mmt_n}건")

        # rom_reference
        ws = wb["rom_reference"]
        rom_n = 0
        for r in iter_rows(ws):
            joint, motion, norm_rom, *_ = r
            if joint and motion:
                conn.execute(text(
                    "INSERT INTO mst_rom_reference(joint,motion,normal_rom) "
                    "VALUES(:j,:m,:nr) ON CONFLICT DO NOTHING"
                ), {"j": joint, "m": motion, "nr": str(norm_rom)})
                rom_n += 1
        print(f"   mst_rom_reference: {rom_n}건")

        # subjective_vocab
        ws = wb["subjective_vocab"]
        sv_n = 0
        for r in iter_rows(ws):
            axis, term_ko, term_en, note, *_ = r
            if axis and term_ko:
                conn.execute(text(
                    "INSERT INTO mst_subjective_vocab(axis,term_ko,term_en,note) "
                    "VALUES(:a,:k,:e,:n) ON CONFLICT DO NOTHING"
                ), {"a": axis, "k": term_ko, "e": term_en, "n": note})
                sv_n += 1
        print(f"   mst_subjective_vocab: {sv_n}건")

        # red_flag
        ws = wb["red_flag"]
        rf_n = 0
        for r in iter_rows(ws):
            cat, sign, action, *_ = r
            if cat and sign:
                conn.execute(text(
                    "INSERT INTO mst_red_flag(category,sign,action) "
                    "VALUES(:c,:s,:a) ON CONFLICT DO NOTHING"
                ), {"c": cat, "s": sign, "a": action})
                rf_n += 1
        print(f"   mst_red_flag: {rf_n}건")

        # imaging_findings  (MRI-척추/MRI-관절 → MRI 로 정규화)
        ws = wb["imaging_findings"]
        img_n = 0
        for r in iter_rows(ws):
            modality, finding, term_en, *_ = r
            if modality and finding:
                # "MRI-척추" → "MRI", "X-ray" → "X-ray" (유지)
                m_str = str(modality)
                modality_norm = m_str if m_str in ("X-ray","MRI","US","CT") else m_str.split("-")[0]
                conn.execute(text(
                    "INSERT INTO mst_imaging_finding(modality,finding,term_en) "
                    "VALUES(:m,:f,:t) ON CONFLICT DO NOTHING"
                ), {"m": modality_norm, "f": finding, "t": term_en})
                img_n += 1
        print(f"   mst_imaging_finding: {img_n}건")

        # grading_scale
        ws = wb["grading_scale"]
        gs_n = 0
        for r in iter_rows(ws):
            if len(r) < 5:
                continue
            scale, target, grades, note, src = r[0], r[1], r[2], r[3], r[4]
            if scale:
                conn.execute(text(
                    "INSERT INTO mst_grading_scale(scale,target,grades,note,src) "
                    "VALUES(:s,:t,:g,:n,:sr) ON CONFLICT DO NOTHING"
                ), {"s": scale, "t": target, "g": grades, "n": note, "sr": src})
                gs_n += 1
        print(f"   mst_grading_scale: {gs_n}건")

        # outcome_scale
        ws = wb["outcome_scale"]
        os_n = 0
        for r in iter_rows(ws):
            if len(r) < 3:
                continue
            scale, target, score_range = r[0], r[1], r[2]
            if scale:
                conn.execute(text(
                    "INSERT INTO mst_outcome_scale(scale,target_region,score_range) "
                    "VALUES(:s,:t,:sr) ON CONFLICT DO NOTHING"
                ), {"s": scale, "t": target, "sr": str(score_range) if score_range else None})
                os_n += 1
        print(f"   mst_outcome_scale: {os_n}건")

        # peripheral_nerve
        ws = wb["peripheral_nerve"]
        pn_n = 0
        for r in iter_rows(ws):
            nerve, site, symptom, exam, rel_dx, *_ = r
            if nerve:
                conn.execute(text(
                    "INSERT INTO mst_peripheral_nerve(nerve,site,symptom,exam,related_dx) "
                    "VALUES(:nv,:s,:sy,:ex,:rd)"
                ), {"nv": nerve, "s": site, "sy": symptom,
                    "ex": exam, "rd": rel_dx})
                pn_n += 1
        print(f"   mst_peripheral_nerve: {pn_n}건")

        # ── 9. map_alias (charting_abbrev) ────────────────────────────────────
        print("▶ 9. map_alias 삽입…")
        # 카테고리 → target_table 매핑
        CAT_TABLE: dict[str, str] = {
            "측성":    "laterality",
            "진단":    "mst_diagnosis",
            "검사":    "mst_exam_item",
            "소견":    "mst_exam_item",
            "영상검사": "mst_exam_item",
            "해부":    "mst_body_part",
            "신경":    "mst_body_part",
            "방향":    "mst_body_part",
            "시술":    "mst_order",
            "재활":    "mst_order",
            "처치":    "mst_order",
            "용법":    "mst_order",
        }
        LATERALITY_MAP = {
            "Lt": "L", "Lt.": "L", "Rt": "R", "Rt.": "R",
            "L": "L", "R": "R", "Bil": "B", "Bilateral": "B",
            "양측": "B", "좌": "L", "우": "R",
        }
        ws = wb["charting_abbrev"]
        alias_n = 0
        for r in iter_rows(ws):
            cat, abbrev, full_term, *_ = r
            if not abbrev or not full_term:
                continue
            abbrev = str(abbrev).strip()
            # 측성 직접 매핑
            if str(cat) == "측성" and abbrev in LATERALITY_MAP:
                conn.execute(text(
                    "INSERT INTO map_alias(surface,canonical_id,target_table,"
                    "alias_type,source,confidence) "
                    "VALUES(:s,:c,'laterality','약어','std',1.0) ON CONFLICT DO NOTHING"
                ), {"s": abbrev, "c": LATERALITY_MAP[abbrev]})
                alias_n += 1
                continue
            tbl = CAT_TABLE.get(str(cat))
            if not tbl:
                continue  # 'general' 등 allowed table 없음 → 스킵
            conn.execute(text(
                "INSERT INTO map_alias(surface,canonical_id,target_table,"
                "alias_type,source,confidence) "
                "VALUES(:s,:c,:t,'약어','std',0.95) ON CONFLICT DO NOTHING"
            ), {"s": abbrev, "c": str(full_term).strip(), "t": tbl})
            alias_n += 1
        print(f"   map_alias: {alias_n}건")

        # ── 10. rel_diagnosis_exam ────────────────────────────────────────────
        print("▶ 10. rel_diagnosis_exam 삽입…")
        existing_dx = {r[0] for r in conn.execute(text("SELECT code FROM diagnosis"))}
        ws = wb["rel_diagnosis_exam"]
        rde_n, rde_skip = 0, 0
        for r in iter_rows(ws):
            dx_name, key_exam, *_ = r
            if not dx_name or not key_exam:
                continue
            dx_code = DX_NAME_TO_CODE.get(str(dx_name).strip())
            if not dx_code or dx_code not in existing_dx:
                print(f"   ⚠  진단 없음 → 스킵: {dx_name!r}")
                rde_skip += 1
                continue
            for token in _split_exam_tokens(str(key_exam)):
                item_id = _resolve_exam(token, exam_name_to_id)
                if item_id:
                    conn.execute(text(
                        "INSERT INTO rel_diagnosis_exam(diagnosis_code,item_id,source) "
                        "VALUES(:dc,:ii,'std') ON CONFLICT DO NOTHING"
                    ), {"dc": dx_code, "ii": item_id})
                    rde_n += 1
                else:
                    if token not in (None, "") and EXAM_ALIASES.get(token) is not None:
                        print(f"   ⚠  검사명 매핑 실패: {token!r}")
        print(f"   rel_diagnosis_exam: {rde_n}건 삽입, {rde_skip}건 스킵")

        # ── 11. rel_structure_exam (exam_item.target_node 기반 자동 등록) ─────
        print("▶ 11. rel_structure_exam 삽입…")
        rows_sn = conn.execute(text(
            "SELECT item_id, target_node FROM mst_exam_item WHERE target_node IS NOT NULL"
        ))
        rse_n = 0
        for item_id, node_id in rows_sn:
            conn.execute(text(
                "INSERT INTO rel_structure_exam(node_id, item_id) "
                "VALUES(:n,:i) ON CONFLICT DO NOTHING"
            ), {"n": node_id, "i": item_id})
            rse_n += 1
        print(f"   rel_structure_exam: {rse_n}건 (target_node 기반)")

        # ── 12. rel_part_diagnosis (지식 기반) ────────────────────────────────
        print("▶ 12. rel_part_diagnosis 삽입…")
        rpd_n, rpd_skip = 0, 0
        for bp_name, dx_code in PART_DIAGNOSIS:
            bp_id = name_to_bp_id.get(bp_name)
            if not bp_id:
                print(f"   ⚠  부위 없음 → 스킵: {bp_name!r}")
                rpd_skip += 1
                continue
            if dx_code not in existing_dx:
                print(f"   ⚠  진단 코드 없음 → 스킵: {dx_code}")
                rpd_skip += 1
                continue
            conn.execute(text(
                "INSERT INTO rel_part_diagnosis(body_part_id,diagnosis_code,support) "
                "VALUES(:b,:d,1.0) ON CONFLICT DO NOTHING"
            ), {"b": bp_id, "d": dx_code})
            rpd_n += 1
        print(f"   rel_part_diagnosis: {rpd_n}건 삽입, {rpd_skip}건 스킵")

    # ── 결과 요약 ────────────────────────────────────────────────────────────
    print("\n✅ 마스터 데이터 적재 완료")


if __name__ == "__main__":
    main()
