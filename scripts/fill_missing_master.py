"""미흡 판정 6개 시트에 누락 데이터를 추가한다.
기존 행은 보존하고 새 행만 append.
실행: python scripts/fill_missing_master.py
"""
from __future__ import annotations
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

SRC = Path("master_data/orthopedic_master_data.xlsx")
DST = SRC  # 원본 덮어쓰기 (백업 먼저 생성)
BAK = SRC.with_suffix(".bak.xlsx")

# ── 백업
shutil.copy2(SRC, BAK)
print(f"백업: {BAK}")

# ── 추가 데이터 정의 ──────────────────────────────────────────────

# 1. diagnosis ─────────────────────────────────────────────────────
# (name_ko, name_en, kcd_icd10, icd11, region, icd11_src, 필요이유)
DIAG_REASON = "정형외과 개인병원 다빈도 진단코드 보강"
new_diagnosis = [
    # ── M25 관절통/강직 계열 (청구량 최상위) ──────────────────────
    ("관절통(부위불명)",       "Joint pain, unspecified",               "M25.50", "FA90",    "전신",  "kb", DIAG_REASON),
    ("어깨 관절통",            "Shoulder pain",                          "M25.51", "FA90.0",  "어깨",  "kb", DIAG_REASON),
    ("팔꿈치 관절통",          "Elbow pain",                             "M25.52", "FA90.1",  "팔꿈치","kb", DIAG_REASON),
    ("손목/손 관절통",         "Wrist and hand pain",                    "M25.54", "FA90.3",  "수부",  "kb", DIAG_REASON),
    ("고관절 관절통",          "Hip pain",                               "M25.55", "FA90.4",  "고관절","kb", DIAG_REASON),
    ("무릎 관절통",            "Knee pain",                              "M25.56", "FA90.5",  "무릎",  "kb", DIAG_REASON),
    ("발목/발 관절통",         "Ankle and foot pain",                    "M25.57", "FA90.6",  "족부",  "kb", DIAG_REASON),
    ("어깨 관절 강직",         "Stiffness of shoulder joint",            "M25.61", "FA91.0",  "어깨",  "kb", DIAG_REASON),
    ("무릎 관절 강직",         "Stiffness of knee joint",                "M25.66", "FA91.5",  "무릎",  "kb", DIAG_REASON),
    ("관절 삼출",              "Effusion of joint",                      "M25.40", "FA93",    "전신",  "kb", DIAG_REASON),

    # ── M10 통풍 ─────────────────────────────────────────────────
    ("통풍(부위불명)",         "Gout, unspecified",                      "M10.00", "5A20",    "전신",  "web","정형외과 개원의 다빈도 — KCD 통풍 기본 코드"),
    ("발목/발 통풍",           "Gout of ankle and foot",                 "M10.07", "5A20",    "족부",  "kb", "발가락(무지) 통풍 발작 다빈도"),
    ("무릎 통풍",              "Gout of knee",                           "M10.06", "5A20",    "무릎",  "kb", DIAG_REASON),
    ("어깨 통풍",              "Gout of shoulder",                       "M10.01", "5A20",    "어깨",  "kb", DIAG_REASON),
    ("손목/손 통풍",           "Gout of wrist and hand",                 "M10.04", "5A20",    "수부",  "kb", DIAG_REASON),

    # ── 신경근병증 세분 ──────────────────────────────────────────
    ("좌골신경통",             "Sciatica",                               "M54.3",  "ME84.1",  "척추",  "web","요추 외래 최다빈도 상병 — HNP 동반 유무 불문"),
    ("좌골신경통 동반 요통",   "Lumbago with sciatica",                  "M54.4",  "ME84.1",  "척추",  "web","임상 최다빈도 조합코드"),
    ("추간판 장애 동반 신경근병증","Disc degeneration with radiculopathy","M51.1",  "FA80",    "척추",  "web","요추 신경근병증 핵심코드 — HNP+방사통"),
    ("경추신경근병증",         "Cervical radiculopathy",                 "M54.12", "FA80",    "척추",  "kb", "M50.2와 청구 구분 필요"),
    ("천장관절 장애",          "Sacroiliac joint disorder",              "M53.3",  "FA84",    "척추",  "kb", "SI joint 기능장애 — 요추통증 감별"),
    ("척수증(경추)",           "Cervical myelopathy",                    "M47.11", "FA84",    "척추",  "kb", "경추 척수압박 — 신경외과 의뢰 기준"),

    # ── M65/M70/M71 건초염·점액낭염 계열 ─────────────────────────
    ("건초염(상세불명)",       "Tenosynovitis, unspecified",             "M65.9",  "FA70",    "전신",  "kb", DIAG_REASON),
    ("방아쇠 수지(엄지)",      "Trigger thumb",                          "M65.30", "FA70.0",  "수부",  "kb", "방아쇠 수지 부위별 세분 — 엄지"),
    ("방아쇠 수지(검지)",      "Trigger index finger",                   "M65.31", "FA70.0",  "수부",  "kb", "방아쇠 수지 부위별 세분 — 검지"),
    ("방아쇠 수지(중지)",      "Trigger middle finger",                  "M65.32", "FA70.0",  "수부",  "kb", "방아쇠 수지 부위별 세분 — 중지"),
    ("방아쇠 수지(약지)",      "Trigger ring finger",                    "M65.33", "FA70.0",  "수부",  "kb", "방아쇠 수지 부위별 세분 — 약지"),
    ("방아쇠 수지(소지)",      "Trigger little finger",                  "M65.34", "FA70.0",  "수부",  "kb", "방아쇠 수지 부위별 세분 — 소지"),
    ("이두건 건초염",          "Bicipital tenosynovitis",                "M65.21", "FA70",    "어깨",  "kb", "상완이두건 장두 건초 — 이두근고랑 압통"),
    ("주두 점액낭염",          "Olecranon bursitis",                     "M70.2",  "FA73",    "팔꿈치","web","팔꿈치 후방 낭 — 외상/반복압박 다빈도"),
    ("슬개전 점액낭염",        "Prepatellar bursitis",                   "M70.4",  "FA73",    "무릎",  "web","무릎 앞 낭 — 바닥작업/종교 활동 관련"),
    ("기타 무릎 점액낭염",     "Other bursitis of knee",                 "M70.5",  "FA73",    "무릎",  "kb", "거위발낭/오금낭 포함"),
    ("대전자 점액낭염",        "Trochanteric bursitis",                  "M70.60", "FA73",    "고관절","web","대전자 외측 통증 — 여성·비만 다빈도"),
    ("견봉하 점액낭염",        "Subacromial bursitis",                   "M75.51", "FA72",    "어깨",  "kb", "충돌증후군 동반 — 어깨 외래 핵심"),
    ("슬와 낭종(Baker's cyst)","Popliteal cyst",                         "M71.20", "FA73",    "무릎",  "web","후방 무릎 낭종 — 반월상 손상 동반 빈번"),

    # ── G57 말초신경 포착 계열 ────────────────────────────────────
    ("이상감각 대퇴통증(LFCN)","Meralgia paraesthetica",                 "G57.2",  "8C10.3",  "고관절","web","외측대퇴피신경 포착 — 서혜부~외측대퇴 저림"),
    ("비골신경병증",           "Lateral popliteal nerve lesion",         "G57.3",  "8C10.5",  "무릎",  "web","무릎 외측 비골두 압박"),
    ("족근관 증후군",          "Tarsal tunnel syndrome",                 "G57.5",  "8C10",    "족부",  "web","내과 후방 경골신경 포착 — 족저 저림"),
    ("Morton 신경종",          "Morton's neuralgia (plantar nerve)",     "G57.6",  "8C10",    "족부",  "web","중족부 지신경 압박 — 발볼 통증·저림"),

    # ── 관절증 세분 ──────────────────────────────────────────────
    ("일차성 고관절증(편측)",  "Primary coxarthrosis, unilateral",       "M16.1",  "FA82.0",  "고관절","kb", "M16 세분 — 편측 고관절 OA"),
    ("이차성 고관절증",        "Secondary coxarthrosis",                 "M16.5",  "FA82.1",  "고관절","kb", "외상·선천성 이후 이차 OA"),
    ("일차성 슬관절증(편측)",  "Primary gonarthrosis, unilateral",       "M17.1",  "FA82.2",  "무릎",  "kb", "M17 세분 — 편측 무릎 OA"),
    ("이차성 슬관절증",        "Secondary gonarthrosis",                 "M17.3",  "FA82.3",  "무릎",  "kb", "외상·반월상 이후 이차 OA"),
    ("손가락 관절증",          "Osteoarthritis of fingers",              "M19.04", "FA82",    "수부",  "kb", "손가락 PIP/DIP 퇴행성 관절염"),
    ("관절증(상세불명)",       "Arthrosis, unspecified",                 "M19.90", "FA82",    "전신",  "kb", "비특이적 관절증 — 청구 빈용"),

    # ── 발뒤꿈치 골극 / 발 변형 ──────────────────────────────────
    ("발뒤꿈치 골극",          "Calcaneal spur",                         "M77.3",  "FA70",    "족부",  "web","족저근막염 동반 다빈도 X-ray 소견"),
    ("무지강직증",             "Hallux rigidus",                         "M20.2",  "FB90",    "족부",  "kb", "1st MTP 관절 강직 — 신발 선택 장애"),
    ("후천성 편평족",          "Acquired flatfoot",                      "M21.60", "FB90",    "족부",  "kb", "후경골건 기능부전 연관"),
    ("요족(pes cavus)",        "Pes cavus",                              "M21.70", "FB90",    "족부",  "kb", "족저근막·아킬레스 연관"),

    # ── 골절 코드 (개원의 응급/F/U) ──────────────────────────────
    ("근위 상완골 골절",       "Fracture of upper end of humerus",       "S42.2",  "NB90",    "어깨",  "kb", "낙상 노인 어깨 다빈도 골절"),
    ("원위 요골 골절(Colles)", "Fracture of lower end of radius",        "S52.5",  "NB93",    "수부",  "kb", "손목 골절 최다빈도"),
    ("대퇴경부 골절",          "Fracture of femoral neck",               "S72.0",  "NB92",    "고관절","kb", "고령 낙상 — 즉시 입원의뢰"),
    ("슬개골 골절",            "Fracture of patella",                    "S82.0",  "NB93",    "무릎",  "kb", "직접충격 — 보존 vs 수술 결정"),
    ("경골 근위부 골절",       "Fracture of upper end of tibia",         "S82.1",  "NB93",    "무릎",  "kb", "Schatzker 분류 적용"),
    ("종골 골절",              "Fracture of calcaneus",                  "S92.0",  "NB93",    "족부",  "kb", "낙상/높이 추락 — CT 필수"),
    ("주상골 골절",            "Fracture of navicular bone of hand",     "S62.0",  "NB93",    "수부",  "kb", "해부학적 코담배갑 압통 — X-ray 음성도 의심"),
    ("척추 압박골절(외상성)",  "Traumatic vertebral fracture",           "S22.0",  "NB92",    "척추",  "kb", "흉요추 이행부 다빈도"),

    # ── 기타 중요 누락 ───────────────────────────────────────────
    ("류마티스 관절염(상세불명)","Rheumatoid arthritis, unspecified",     "M06.9",  "FA20",    "전신",  "kb", "정형외과 초진 감별 — 류마티스 의뢰 기준"),
    ("대퇴골두 무혈성 괴사",   "Avascular necrosis of femoral head",     "M87.00", "FB81",    "고관절","web","스테로이드/알코올 연관 — MRI 조기 진단"),
    ("박리성 골연골염",        "Osteochondritis dissecans",              "M93.20", "FA90",    "무릎",  "kb", "청소년 무릎/발목 — MRI 필수"),
    ("신경통(상세불명)",       "Neuralgia, unspecified",                  "M79.2",  "8C10",    "전신",  "kb", "신경병성 통증 비특이적 — 대상포진후 포함"),
    ("기타 척추증",            "Other spondylosis",                      "M47.8",  "FA84",    "척추",  "kb", "M47.1/M47.2 외 척추증"),
    ("흉추 추간판 장애",       "Thoracic disc disorder",                 "M51.4",  "FA80",    "척추",  "kb", "흉추 디스크 — 드물지만 흉통 감별 필수"),
    ("척추 불안정증",          "Instability of spine",                   "M53.2",  "FA84",    "척추",  "kb", "전방전위증 등 역학적 불안정"),
    ("수술후 척추 증후군",     "Postlaminectomy syndrome",               "M96.1",  "FA84",    "척추",  "kb", "수술 후 F/U 환자 재내원"),
    ("복합부위통증증후군(CRPS)","Complex regional pain syndrome",         "M89.0",  "MG31",    "전신",  "web","외상/수술 후 자율신경 이상 통증"),
    ("근막통증증후군(MPS)",    "Myofascial pain syndrome",               "M79.1",  "MG30",    "전신",  "kb", "TPI 대상 — 압통점 주사치료"),
    ("섬유근통",               "Fibromyalgia",                           "M79.7",  "MG30.01", "전신",  "web","광범위 근골격 통증 — 다학제 접근"),
    ("골다공증성 골절",        "Osteoporotic fracture",                  "M80.0",  "FB83",    "척추",  "kb", "척추 압박골절 원인"),
    ("고관절 충돌증후군(FAI)", "Femoroacetabular impingement",           "M24.85", "FA82",    "고관절","web","젊은 활동성 환자 고관절 통증 — MRI/관절경"),
]

# 2. order_type ────────────────────────────────────────────────────
# (order_type, name_en, description, example, 필요이유)
ORD_REASON = "정형외과 개인병원 비급여·시술 처방 보강"
new_order_type = [
    ("도수치료",       "Manual therapy",             "척추·관절 도수 교정, 근막이완",            "경추 도수치료 6회",           "개원의 비급여 매출 1위 — 물리치료와 별도 코드 필수"),
    ("운동치료",       "Therapeutic exercise",       "코어강화, 스트레칭 등 치료적 운동",        "요추 안정화 운동 8회",         "재활 프로그램 처방 — 도수치료·물리치료와 구분"),
    ("PRP 주사",       "PRP injection",              "자가혈소판풍부혈장 채혈·원심분리·주사",    "무릎 PRP 3회",                "건병증·관절염 재생치료 — 비급여 고수익 시술"),
    ("프롤로테라피",   "Prolotherapy",               "인대·건 부착부 고농도 포도당 증식주사",    "SI joint prolotherapy",       "인대 이완 치료 — 비급여 시술"),
    ("PDRN 주사",      "PDRN injection",             "폴리뉴클레오타이드(DNA) 재생주사",        "아킬레스건 PDRN 4회",         "건병증·연골손상 재생 — 연어주사 계열"),
    ("고주파 신경차단","RF ablation / RFA",          "후관절·천장관절 내측지 고주파 열절제술",   "요추 MBB후 RFA",              "MBB와 별도 — 영상유도 고주파 신경절제"),
    ("TPI",            "Trigger point injection",    "근막통 유발점 국소마취제·스테로이드 주사", "상부 승모근 TPI",             "MPS 핵심 치료 — map_alias에만 있던 약어 처방화"),
    ("점액낭 주사",    "Bursa injection",            "점액낭 내 스테로이드/국소마취제 주사",     "견봉하 bursa injection",      "관절강내주사와 해부학적 위치 구분 필요"),
    ("관절 흡인술",    "Joint aspiration",           "관절액 흡인(진단·치료 목적)",             "무릎 삼출 aspiration",         "삼출 동반 관절염·혈관절증 처치"),
    ("신경성형술",     "Epiduroscopy / Neuroplasty", "경막외강 카테터 신경근 유착 박리술",       "L4-5 신경성형술",             "협착증·수술후유착 — 개원 시술 가능"),
    ("단하지석고",     "Short leg cast (SLSC)",      "발목 이하 고정 — 발목·족부 골절",         "SLSC 6주",                    "캐스트 세분화 — 보험 청구 코드 구분"),
    ("단상지석고",     "Short arm cast (SASC)",      "손목 이하 고정 — 요골/주상골 골절",       "SASC 6주",                    "캐스트 세분화 — 보험 청구 코드 구분"),
    ("장하지석고",     "Long leg cast (LLSC)",       "슬관절 포함 대퇴~발목 고정",              "LLSC 4주",                    "경골 근위부 골절 등 무릎 포함 고정"),
    ("테이핑",         "Kinesio taping",             "근육·인대 지지 키네지오 테이핑",          "슬개건 taping",               "외래 처방 — 물리치료 세부 or 단독"),
    ("체외충격파(집속형)","Focused ESWT",            "집속형 체외충격파 — 건병증·석회화",       "회전근개 석회화 ESWT 3회",    "비급여 시술 — 방사형과 청구 구분"),
    ("초음파유도 시술","US-guided procedure",        "초음파 실시간 유도 주사/흡인/시술",        "US-guided PRP injection",     "시술 정확도·안전성 명시 — 보험 가산"),
]

# 3. medication ────────────────────────────────────────────────────
# (class_계열, ingredient_성분, brand_예시, 필요이유)
MED_REASON_ANAL = "정형외과 1차 진통제 — NSAID 금기·병용"
MED_REASON_LOCAL = "주사 시술 필수 동반 — 신경차단·TPI·관절주사 시 반드시 사용"
MED_REASON_OSTEO = "골다공증 진단 있는데 치료약 없는 모순 해소"
new_medication = [
    # 진통제
    ("진통제(비마약성)",    "아세트아미노펜",                  "타이레놀ER, 써스펜8시간",    MED_REASON_ANAL),
    ("진통제(약한opioid)", "트라마돌",                        "트리돌, 마이올서방정",        "중등도 통증 — NSAID 불충분 시 2차"),
    ("진통제 복합제",       "트라마돌+아세트아미노펜",         "울트라셋, 트리돌에이",        "정형외과 외래 복합 진통제 최다처방"),
    ("진통제 복합제",       "아세트아미노펜+이부프로펜",       "애드빌듀얼액션",              "복합 OTC 급성 통증"),
    # NSAID 추가
    ("NSAID",               "록소프로펜",                     "록소닌, 비록소",              "일본계 NSAID — 국내 개원 다빈도"),
    ("NSAID",               "에토리콕시브",                   "아르콕시아60mg/90mg",         "COX-2 선택적 — 고위험 위장 환자"),
    ("NSAID",               "덱시부프로펜",                   "이지엔6, 덱시부",             "이부프로펜 활성형 — 국내 다빈도"),
    ("NSAID",               "케토프로펜",                     "케토신, 오루스",              "급성 염증 — 외용제 병용"),
    ("NSAID(외용)",         "케토프로펜 외용",                "케토탑패취, 케펜텍겔",        "국소 외용 NSAID — 패취/겔"),
    ("NSAID(외용)",         "디클로페낙 외용",                "볼타렌겔, 디클로페낙겔",      "국소 외용 — 무릎·어깨 관절"),
    # 스테로이드 경구
    ("스테로이드(경구)",    "프레드니솔론",                   "소론도정, 프레드니정",        "단기 경구 스테로이드 — 급성 염증 burst"),
    ("스테로이드(경구)",    "메틸프레드니솔론(경구)",         "메드롤정",                    "어깨·척추 급성기 경구 스테로이드"),
    # 국소마취제 (시술 필수)
    ("국소마취제",          "리도카인",                       "리도카인 앰플 1%/2%",         MED_REASON_LOCAL),
    ("국소마취제",          "부피바카인",                     "마카인 0.25%/0.5%",           "신경차단·ESI·경막외강 장시간 차단"),
    ("국소마취제",          "로피바카인",                     "나로핀 0.2%/0.75%",           "장시간 신경차단 — 마카인 대체"),
    ("국소마취제",          "메피바카인",                     "카르보카인",                  "치과·소수술 국소마취"),
    # 히알루론산 제품명 세분
    ("히알루론산 주사",     "히알루론산(시노비안계열)",       "시노비안, 뉴론, 이오판",      "무릎·고관절 관절강내 — 3회 주사형"),
    ("히알루론산 주사",     "히알루론산(오스테닐계열)",       "오스테닐, 듀로란",            "단회 주사형 히알루론산"),
    ("히알루론산 주사",     "히알루론산(히루안계열)",         "히루안플러스",                "국내 최초 히알루론산 — 보령"),
    # PDRN
    ("재생주사",            "PDRN(폴리뉴클레오타이드)",       "리쥬란힐러, 플라센텍스",      "건병증·연골 재생 — 연어DNA 유래"),
    # 항골다공증
    ("항골다공증(경구)",    "알렌드로네이트",                 "포사맥스, 맥스마빌, 오스포스", MED_REASON_OSTEO),
    ("항골다공증(경구)",    "리세드로네이트",                 "악토넬, 오스테론",            MED_REASON_OSTEO),
    ("항골다공증(IV)",      "졸레드론산",                     "조메타, 아클라스타",          "연 1회 정맥주사 — 경구 불가 환자"),
    ("항골다공증(보조제)",  "칼슘+비타민D",                   "오스칼, 칼로퍼, 디카맥스",    "골다공증 기본 보조 — 전 환자 병용"),
    ("항골다공증(주사)",    "데노수맙",                       "프롤리아 60mg",               "6개월 1회 피하 — 중증 골다공증"),
    # 신경보호·통증보조
    ("비타민B12",           "메코발라민",                     "메코정, 메티코발 주사",       "신경병성 통증 보조 — 말초신경 보호"),
    ("신경병성통증",        "아미트리프틸린",                 "에트라빌, 에나폰정",          "만성통증 저용량·수면장애 동반"),
    ("신경병성통증",        "벤라팍신",                       "이펙사XR, 벤라민",            "둘록세틴 불내성 대체"),
    # 통풍
    ("통풍치료",            "콜히친",                         "콜킨정",                      "통풍 급성기 필수 — 24h 내 투여"),
    ("통풍치료",            "알로푸리놀",                     "자이로릭, 알로퓨린",          "통풍 요산저하 유지치료"),
    ("통풍치료",            "페북소스타트",                   "페브릭정",                    "알로푸리놀 불내성 대체"),
    # 근이완제 추가
    ("근이완제",            "티자니딘",                       "시르달루드, 티자렉스",        "경련성 근긴장 — 척추 경련 적합"),
    ("근이완제",            "메토카르바몰",                   "마이오뷰정",                  "급성 근육통·요통 단기 사용"),
]

# 4. imaging_findings ──────────────────────────────────────────────
# (modality, finding_소견, term_en, 필요이유)
IMG_REASON = "정형외과 개인병원 영상판독의뢰서·차팅 표준화"
new_imaging = [
    # CT 신규 섹션
    ("CT", "척추관 직경 협소(mm)",          "Spinal canal stenosis (diameter, mm)",       IMG_REASON),
    ("CT", "신경공 협착",                   "Foraminal stenosis",                          IMG_REASON),
    ("CT", "측방 함요부 협착",              "Lateral recess stenosis",                     IMG_REASON),
    ("CT", "골절 분쇄도(분쇄/비분쇄)",     "Comminuted vs non-comminuted fracture",       IMG_REASON),
    ("CT", "골절편 전위(mm)",               "Fragment displacement (mm)",                  IMG_REASON),
    ("CT", "골유합 진행(callus 형성)",      "Callus formation / bone bridging",            IMG_REASON),
    ("CT", "석회화건염 Gärtner 분류(A/B/C)","Calcific tendinitis Gärtner type A/B/C",     "석회화 치료방침 — ESWT/흡인 결정 기준"),
    ("CT", "황인대 비후(mm)",               "Ligamentum flavum hypertrophy (mm)",          "협착증 수술의뢰 정량 기준"),
    ("CT", "후관절 비대/골극",              "Facet joint hypertrophy / osteophyte",        "협착 기여 인자 — MBB 대상 결정"),
    ("CT", "3D 재건 소견",                  "3D CT reconstruction findings",               "복잡 골절 수술 계획"),
    # MRI 추가 소견
    ("MRI", "추간판 형태(팽륜/탈출/탈출수)/extrusion/sequestration", "Disc bulge / protrusion / extrusion / sequestration", "HNP 세부 형태 — 수술 적응증 판단"),
    ("MRI", "Modic change (Type I/II/III)", "Modic change Type I/II/III",                  "종판 변성 — 만성통증 예후"),
    ("MRI", "척추관협착 등급(Schizas A-D)", "Stenosis grade Schizas A/B/C/D",             "협착 정도 표준 등급 — 수술의뢰 기준"),
    ("MRI", "척수 T2 신호 변화",            "T2 signal change in spinal cord",             "척수병증 조기 소견 — 경추 MRI 필수"),
    ("MRI", "회전근개 지방변성(Goutallier 0-4)", "Goutallier grade 0-4 fatty infiltration", "수술 예후 — 4등급 이상 재건 불량"),
    ("MRI", "골수 부종(BME)",               "Bone marrow edema (BME)",                     "급성 골절·스트레스반응·AVN 초기"),
    ("MRI", "Facet joint OA(Pathria grade)","Facet joint OA Pathria grade I-III",          "후관절 퇴행 정량 — MBB 적응증"),
    ("MRI", "관절와순 파열(전/후/상/전체)", "Labral tear type (anterior/posterior/superior/full)", "SLAP·Bankart 구분"),
    ("MRI", "반월상연골 파열 유형(종축/수평/방사)", "Meniscal tear type (vertical/horizontal/radial)", "수술 vs 보존 결정"),
    # US 추가 소견
    ("US",  "활액막 비후(mm)",              "Synovial thickening (mm)",                    "관절염 모니터링 — 주사 필요성 판단"),
    ("US",  "신경 단면적(CSA mm²)",         "Nerve cross-sectional area (CSA, mm²)",       "수근관/척골신경 포착 진단 기준"),
    ("US",  "건 두께(mm)",                  "Tendon thickness (mm)",                       "아킬레스건·슬개건 정량 — 건병증 추적"),
    ("US",  "동적 불안정성 소견",           "Dynamic instability on stress US",            "ATFL 동적 발목 불안정성 평가"),
    ("US",  "석회화 크기(mm)·에코양상",    "Calcification size (mm) and echogenicity",    "석회화 흡인 적응증 판단"),
]

# 5. red_flag ──────────────────────────────────────────────────────
# (category, sign_징후, action_조치, 필요이유)
RF_REASON = "외래 오진 방지 — 비근골격계 연관통 감별"
new_red_flag = [
    ("심혈관",      "좌측 어깨/팔 통증+흉압박감+발한·구역 (ACS pattern)",
                    "즉시 ECG·트로포닌·119 이송 — 회전근개와 감별 필수",
                    "심근경색 오진 의료사고 최다보고 — 어깨통증 외래 1순위 감별"),
    ("심혈관",      "찢어지는 등/흉부 통증+양쪽 혈압 차이 (대동맥박리)",
                    "응급 CT 혈관조영+즉시 이송",
                    "흉추통증 감별 — 혈압 측정 양쪽 필수"),
    ("혈전",        "종아리 편측 발적·부종·열감·압통 (DVT 의심)",
                    "D-dimer 채혈+Duplex US 의뢰 — 즉시 항응고 고려",
                    "종아리 통증 외래 — 근육통과 감별 / PE 전단계"),
    ("혈관",        "보행시 종아리 통증+발목 ABI 감소 (PAD/말초혈관)",
                    "혈관외과 의뢰+ABI 측정",
                    "신경인성 파행 vs 혈관성 파행 감별 필수"),
    ("부인과",      "하복부/골반/요추 통증+임신가능 여성+성기출혈 (자궁외임신)",
                    "즉시 부인과 응급의뢰 — 복강내 출혈 위험",
                    "요추/고관절 통증으로 내원한 가임기 여성 감별"),
    ("비뇨기",      "측복부·서혜부 방사통+혈뇨+구역 (요로결석/신우신염)",
                    "비뇨기과 의뢰+소변검사",
                    "요추통증과 혼동 — 측방 요추통 감별"),
    ("종양",        "어깨~팔 통증+Horner 증후군+체중감소 (Pancoast 종양)",
                    "흉부 X-ray 및 CT+종양내과 의뢰",
                    "C8-T1 신경근병증 감별 — 폐첨부암"),
    ("감염",        "관절 발적·열감·발열+단관절 급성 부종 (화농성 관절염)",
                    "즉시 관절천자 세균배양+응급 정형외과/입원 의뢰",
                    "관절통 외래 — 48h 내 진단 못하면 관절 영구손상"),
]

# 6. subjective_vocab ──────────────────────────────────────────────
# (axis, term_ko, term_en, note, 필요이유)
SV_REASON = "SOCRATES 표준 문진 axis 확장 — S섹션 구조화 완성"
new_subj = [
    # axis: radiation (방사)
    ("radiation", "방사 없음",          "No radiation",             "국소 통증",                         SV_REASON),
    ("radiation", "신경근 분포 따름",   "Dermatomal radiation",     "팔/다리 쪽으로 전기처럼 내려감",     SV_REASON),
    ("radiation", "비피부분절성 방사",  "Non-dermatomal referred",  "관절·근막 연관통 패턴",               SV_REASON),
    ("radiation", "양측 방사",          "Bilateral radiation",      "양쪽 동시 내려감",                   SV_REASON),
    ("radiation", "방사 끝점(어디까지)","Radiation endpoint",        "발끝까지/무릎까지/발목까지 등",       SV_REASON),

    # axis: aggravating (악화인자)
    ("aggravating", "앉아 있을 때 악화",      "Worse with sitting",        "요추 추간판/좌골신경통",    SV_REASON),
    ("aggravating", "서 있을 때 악화",        "Worse with standing",       "척추관협착/정맥성",         SV_REASON),
    ("aggravating", "걸을 때 악화",           "Worse with walking",        "신경성 파행/혈관성 파행",   SV_REASON),
    ("aggravating", "앞으로 굽힐 때 악화",    "Worse with forward flexion","HNP/추간판성 통증",         SV_REASON),
    ("aggravating", "뒤로 젖힐 때 악화",      "Worse with extension",      "후관절·협착·분리증",        SV_REASON),
    ("aggravating", "기침·재채기 시 악화",    "Worse with cough/sneeze",   "Valsalva → 추간판 내압↑",   SV_REASON),
    ("aggravating", "무거운 것 들 때 악화",   "Worse with lifting",        "추간판·회전근개 부하",      SV_REASON),
    ("aggravating", "특정 동작(overhead) 악화","Worse with overhead activity","충돌증후군·회전근개",   SV_REASON),

    # axis: relieving (완화인자)
    ("relieving",   "안정/휴식 시 완화",      "Relieved by rest",          "기계적 통증",               SV_REASON),
    ("relieving",   "온찜질 시 완화",         "Relieved by heat",          "근경련·만성 근육통",        SV_REASON),
    ("relieving",   "냉찜질 시 완화",         "Relieved by ice",           "급성 염증·외상 초기",       SV_REASON),
    ("relieving",   "자세 변경 시 완화",      "Relieved by position change","협착증(구부릴 때 완화)",   SV_REASON),
    ("relieving",   "진통제 복용 후 완화",    "Relieved by analgesics",    "약물 반응성 확인",          SV_REASON),
    ("relieving",   "완화 없음(지속통)",      "No relief",                 "Red flag / 염증성",         SV_REASON),

    # axis: time_pattern (시간패턴)
    ("time_pattern","야간통(수면 방해)",       "Nocturnal pain",            "Red flag — 악성종양/염증성", SV_REASON),
    ("time_pattern","아침 강직 > 1시간",       "Morning stiffness >1hr",    "강직성척추염 등 염증성",    SV_REASON),
    ("time_pattern","아침 강직 < 30분",        "Morning stiffness <30min",  "퇴행성 관절염",             SV_REASON),
    ("time_pattern","활동 후 악화·안정 시 완화","Worse after activity",      "OA·기계적 통증 패턴",      SV_REASON),
    ("time_pattern","안정 시에도 지속",        "Pain at rest",              "염증성·신경병성",           SV_REASON),
    ("time_pattern","지속성 통증",             "Constant pain",             "vs 간헐성 구분",            SV_REASON),
    ("time_pattern","간헐성 통증",             "Intermittent pain",         "유발인자 관련",             SV_REASON),
    ("time_pattern","점진적 악화",             "Progressive worsening",     "Red flag 확인 필요",        SV_REASON),

    # axis: associated_symptoms (동반증상)
    ("associated",  "저림/감각이상",           "Paresthesia / numbness",    "신경근·말초신경 포착",      SV_REASON),
    ("associated",  "근력 약화",               "Muscle weakness",           "신경학적 손상 — MMT 필요",  SV_REASON),
    ("associated",  "배뇨·배변 장애",          "Bladder / bowel dysfunction","마미증후군 — 즉시 Red flag", SV_REASON),
    ("associated",  "부종(swelling)",          "Swelling / edema",          "외상·염증·혈전",            SV_REASON),
    ("associated",  "관절 잠김(locking)",      "Locking",                   "반월상연골·유리체 감별",    SV_REASON),
    ("associated",  "관절 불안정감(giving way)","Giving way / instability",  "인대 손상·반월상 연관",     SV_REASON),
    ("associated",  "발열·오한",               "Fever / chills",            "감염성 관절염 — Red flag",  SV_REASON),
    ("associated",  "체중 감소",               "Unexplained weight loss",   "악성종양 Red flag",         SV_REASON),

    # axis: severity_impact (중증도·일상 영향)
    ("severity",    "수면 방해",               "Sleep disturbance",         "야간통과 연계",             SV_REASON),
    ("severity",    "보행 거리 제한(m)",        "Walking distance limitation","신경인성 파행 정량",       SV_REASON),
    ("severity",    "계단 오르내리기 제한",     "Stair climbing limitation", "무릎·고관절 기능 지표",     SV_REASON),
    ("severity",    "업무·직업 기능 저하",      "Work/occupational impairment","장해 판정·산재 연계",    SV_REASON),
    ("severity",    "ADL 제한(옷입기/씻기)",    "ADL limitation",            "중증도 기능 평가",          SV_REASON),

    # quality 추가 (기존 axis에 추가)
    ("quality",     "경련/쥐남",               "Cramping",                  "근경련·혈관성",             SV_REASON),
    ("quality",     "욱신거림/박동성",          "Throbbing / pulsating",     "혈관성·감염성 감별",        SV_REASON),
    ("quality",     "깊은/뼈까지 아픈",         "Deep / bone pain",          "골수염·종양·골절 시사",     SV_REASON),
    ("quality",     "압박감",                   "Pressure / heaviness",      "심혈관 감별 주의",          SV_REASON),
]


# ── Excel 업데이트 ────────────────────────────────────────────────

def append_rows(wb, sheet_name: str, columns: list, new_rows: list[tuple]):
    ws = wb[sheet_name]
    before = ws.max_row
    for row_data in new_rows:
        ws.append(list(row_data))
    after = ws.max_row
    print(f"  {sheet_name}: {before-1}행 → {after-1}행 (+{after-before}행)")

wb = load_workbook(SRC)

print("\n[추가 작업 시작]")
append_rows(wb, "diagnosis",        ["name_ko","name_en","kcd_icd10","icd11","region","icd11_src","필요이유"], new_diagnosis)
append_rows(wb, "order_type",       ["order_type","name_en","description","example","필요이유"], new_order_type)
append_rows(wb, "medication",       ["class_계열","ingredient_성분","brand_예시","필요이유"], new_medication)
append_rows(wb, "imaging_findings", ["modality","finding_소견","term_en","필요이유"], new_imaging)
append_rows(wb, "red_flag",         ["category","sign_징후","action_조치","필요이유"], new_red_flag)
append_rows(wb, "subjective_vocab", ["axis","term_ko","term_en","note","필요이유"], new_subj)

wb.save(DST)
print(f"\n[저장 완료] {DST}")
