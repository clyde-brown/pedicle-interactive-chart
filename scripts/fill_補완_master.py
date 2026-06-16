"""보완 필요 8개 시트에 누락 데이터를 추가/수정한다.
기존 행은 보존. rel_diagnosis_exam의 잘못된 매핑 1건 수정 포함.
실행: python scripts/fill_補완_master.py
"""
from __future__ import annotations
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

SRC = Path("master_data/orthopedic_master_data.xlsx")
BAK = SRC.with_suffix(".bak2.xlsx")
shutil.copy2(SRC, BAK)
print(f"백업: {BAK}\n")

# ══════════════════════════════════════════════════════════════════
# 1. rom_reference  — 흉추·손가락·TMJ
# ══════════════════════════════════════════════════════════════════
# (joint, motion, normal_ROM, 필요이유)
ROM_R = "정상 가동범위 기준 보강 — 흉추/손가락/TMJ"
new_rom = [
    ("흉추",   "굴곡/신전",          "20~45° / 25~45°",    ROM_R),
    ("흉추",   "측방굴곡/회전",      "20~40° / 35~50°",    ROM_R),
    ("손가락", "MCP 굴곡/신전",      "90° / 45°",          "방아쇠수지·수지골절 F/U 필수"),
    ("손가락", "PIP 굴곡/신전",      "100° / 0°",          "방아쇠수지·수지골절 F/U 필수"),
    ("손가락", "DIP 굴곡/신전",      "90° / 0°",           "방아쇠수지·수지골절 F/U 필수"),
    ("엄지",   "CMC 외전/대립",      "70° / 기능평가",      "드퀘르벵·CMC OA 추적"),
    ("TMJ",    "개구/전방돌출/측방이동", "40~50mm / 8~12mm / 8~12mm", "TMJ 진찰 기준값 — body_part·exam_item과 정합"),
]

# ══════════════════════════════════════════════════════════════════
# 2. grading_scale  — Eaton-Littler / Rockwood / Denis / Goutallier / Tonnis
# ══════════════════════════════════════════════════════════════════
# (scale, 적용, grades, 필요이유, src)
new_grading = [
    ("Eaton-Littler",   "엄지 CMC관절 OA",         "I~IV",   "중장년 여성 외래 다빈도 — CMC 치료방침 결정", "web"),
    ("Rockwood",        "AC joint 탈리",            "I~VI",   "어깨 낙상/스포츠 외상 — I-III 보존, IV-VI 수술", "web"),
    ("Denis",           "흉요추 척추골절",          "A~D(4형)","골다공증성 압박골절 type A/B/C/D — 수술 여부", "web"),
    ("Goutallier",      "회전근개 지방변성(MRI/CT)","0~4",    "수술 예후 — 3등급↑ 재건 불량 기준",            "web"),
    ("Tonnis",          "고관절 OA(X-ray)",         "0~3",    "K-L 대안 — 고관절 OA 수술의뢰 기준",           "web"),
    ("Ficat & Arlet",   "대퇴골두 무혈성괴사",      "0~IV",   "AVN 외래 경과 추적 — MRI 기반",                 "web"),
    ("Snyder SLAP",     "관절와순 SLAP 병변",       "I~IV",   "MRI 판독 소견 기술 — SLAP 수술 적응증",         "web"),
]

# ══════════════════════════════════════════════════════════════════
# 3. outcome_scale  — EQ-5D / PRWE / LEFS / PSFS
# ══════════════════════════════════════════════════════════════════
# (scale, 적용부위, 범위, 필요이유)
new_outcome = [
    ("EQ-5D-5L",                   "전신(삶의 질)",  "0~1(index) / VAS 0~100", "장애진단서·보험사 제출·임상연구 국제표준 — 누락 시 행정 불가"),
    ("PRWE (Patient-Rated Wrist)", "손목",           "0~100",                  "드퀘르벵·TFCC·Colles 경과추적 — DASH보다 손목 특이적"),
    ("LEFS (Lower Extremity Func)","하지 전체",      "0~80",                   "무릎·발목·고관절 광범위 하지기능 — PT 협진 의뢰 공문 표준"),
    ("PSFS (Patient Specific Func)","환자 개인목표", "0~10(3항목)",            "5분 외래용 개인화 기능 추적 — 환자 스스로 목표 설정"),
    ("GRC (Global Rating of Change)","전반적 변화",  "-7 ~ +7",                "치료 효과 주관적 추적 — 단일 문항"),
    ("AOFAS Ankle-Hindfoot",       "발목/후족부",    "0~100",                  "FAAM과 병용 — 국내 논문 표준척도"),
]

# ══════════════════════════════════════════════════════════════════
# 4. peripheral_nerve  — TOS / AIN / PIN / 장흉신경 / 견갑상신경
# ══════════════════════════════════════════════════════════════════
# (nerve, site, symptom, exam, 관련진단, 필요이유)
new_nerve = [
    ("흉곽출구 신경혈관속(TOS)",
     "전사각근/늑쇄공간/오훼돌기하",
     "팔 저림·피로감·맥박감소",
     "Roos test(EAST), Adson test, Wright test",
     "흉곽출구증후군(TOS)",
     "경추신경근병증·회전근개와 감별 — Roos 민감도 84%"),
    ("전방골간신경(AIN, 정중신경 분지)",
     "전완 근위부(Frohse arcade 근위)",
     "OK sign 불가·엄지-검지 말단굴곡 마비",
     "Pinch grip test(OK sign 실패)",
     "AIN 증후군",
     "손목 굴곡력 약화 감별 — 수근관증후군과 구분"),
    ("후방골간신경(PIN, 요골신경 심부분지)",
     "회외근(Arcade of Frohse)",
     "손목하수(wrist drop)·손등 감각저하",
     "저항 회외전, 저항 중지신전",
     "PIN 증후군(회외근 증후군)",
     "외측상과염 감별 — 요골신경 압박"),
    ("장흉신경(Long thoracic nerve)",
     "전거근 신경지배",
     "견갑골 익상(scapular winging)",
     "Wall push-up — 견갑골 내측연 돌출",
     "전거근 마비",
     "어깨 통증 원인 — 익상견갑 조기 발견"),
    ("견갑상신경(Suprascapular nerve)",
     "견갑절흔/오훼상완인대",
     "극상·극하근 위축·어깨 심부통",
     "견갑극 두드림, 극하근 저항 외회전 약화",
     "견갑상신경 포착",
     "회전근개파열 감별 — 위축 심하면 신경 손상 의심"),
]

# ══════════════════════════════════════════════════════════════════
# 5. charting_abbrev  — ~30개 추가
# ══════════════════════════════════════════════════════════════════
# (category, abbrev, full_term, korean, 필요이유)
CA_R = "실무 차팅 약어 보강 — LLM 추출 정확도"
new_abbrev = [
    # 영상 검사
    ("영상검사",  "CT",      "Computed Tomography",                "컴퓨터단층촬영",      CA_R),
    ("영상검사",  "EMG",     "Electromyography",                   "근전도검사",          CA_R),
    ("영상검사",  "NCS",     "Nerve Conduction Study",             "신경전도검사",        CA_R),
    ("영상검사",  "BMD",     "Bone Mineral Density",               "골밀도",              CA_R),
    ("영상검사",  "DXA",     "Dual-energy X-ray Absorptiometry",   "이중에너지 X선흡수",  CA_R),
    ("영상검사",  "W/B",     "Weight-bearing",                     "체중부하 촬영",       CA_R),
    ("영상검사",  "AP/Lat",  "Anteroposterior / Lateral",          "정면/측면",           CA_R),
    # 진단
    ("진단",      "AVN",     "Avascular Necrosis",                 "무혈성 괴사",         CA_R),
    ("진단",      "AS",      "Ankylosing Spondylitis",             "강직성 척추염",       CA_R),
    ("진단",      "RA",      "Rheumatoid Arthritis",               "류마티스 관절염",     CA_R),
    ("진단",      "MPS",     "Myofascial Pain Syndrome",           "근막통증증후군",      CA_R),
    ("진단",      "TOS",     "Thoracic Outlet Syndrome",           "흉곽출구증후군",      CA_R),
    ("진단",      "CTS",     "Carpal Tunnel Syndrome",             "수근관증후군",        CA_R),
    ("진단",      "FAI",     "Femoroacetabular Impingement",       "대퇴비구충돌증후군",  CA_R),
    ("진단",      "ITB",     "Iliotibial Band",                    "장경인대",            CA_R),
    ("진단",      "GTPS",    "Greater Trochanteric Pain Syndrome", "대전자통증증후군",    CA_R),
    ("진단",      "PHN",     "Postherpetic Neuralgia",             "대상포진후신경통",    CA_R),
    ("진단",      "CRPS",    "Complex Regional Pain Syndrome",     "복합부위통증증후군",  CA_R),
    # 시술/처치
    ("시술",      "ESWT",    "Extracorporeal Shock Wave Therapy",  "체외충격파치료",      CA_R),
    ("시술",      "PDRN",    "Polydeoxyribonucleotide",            "폴리뉴클레오타이드(연어주사)", CA_R),
    ("시술",      "HA",      "Hyaluronic Acid injection",          "히알루론산 주사",     CA_R),
    ("시술",      "RFA",     "Radiofrequency Ablation",            "고주파 신경절제술",   CA_R),
    ("시술",      "PRF",     "Pulsed Radiofrequency",              "고주파 신경조절술",   CA_R),
    ("시술",      "SIJ inj", "Sacroiliac Joint injection",         "천장관절 주사",       CA_R),
    ("시술",      "US-guided","Ultrasound-guided",                 "초음파 유도하",       CA_R),
    # 처방/기록 일반
    ("일반",      "R/O",     "Rule out",                           "감별진단",            CA_R),
    ("일반",      "WNL",     "Within Normal Limits",               "정상범위",            CA_R),
    ("일반",      "ADL",     "Activities of Daily Living",         "일상생활동작",        CA_R),
    ("일반",      "RTW",     "Return to Work",                     "업무복귀",            CA_R),
    ("일반",      "PRN",     "pro re nata",                        "필요시",              CA_R),
    ("일반",      "c/o",     "complains of",                       "주소",                CA_R),
    # 해부/구조
    ("해부",      "QL",      "Quadratus Lumborum",                 "요방형근",            CA_R),
    ("해부",      "HS",      "Hamstring",                          "슬굴근",              CA_R),
    ("해부",      "PF",      "Plantar Fascia",                     "족저근막",            CA_R),
    ("해부",      "AT",      "Achilles Tendon",                    "아킬레스건",          CA_R),
    ("해부",      "CFT",     "Calcific Tendinitis",                "석회화건염",          CA_R),
]

# ══════════════════════════════════════════════════════════════════
# 6. exam_item  — 고관절 special test + 어깨 + 팔꿈치 + 척추
# ══════════════════════════════════════════════════════════════════
# (category, name, name_full, target_part, target_structure, value_type, tests_for, 필요이유)
EX_R = "O섹션 이학적 검사 구조화 — 보완 필요 누락항목"
new_exam = [
    ("special_test", "Trendelenburg test",
     "Trendelenburg sign",
     "고관절", "중둔근/상둔신경", "+/-",
     "중둔근 약화, 상둔신경 병변",
     "외래 매일 시행 — 보행분석·중둔근TPI 전 필수"),
    ("special_test", "Thomas test",
     "Thomas hip flexion contracture test",
     "고관절", "장요근", "각도/각도",
     "고관절 굴곡 구축, 장요근 단축",
     "고관절OA/전치환 전 필수 — 구축 각도 기록"),
    ("special_test", "Ober test",
     "Ober's IT band tightness test",
     "고관절", "장경인대/TFL", "+/-",
     "IT band 구축, TFL 단축",
     "대전자통증증후군·ITBS 감별"),
    ("special_test", "Neer sign",
     "Neer impingement sign",
     "어깨", "견봉하공간/극상근", "+/-",
     "견봉하 충돌증후군(impingement)",
     "충돌증후군 선별 — 민감도 78%, 음성예측도 96%"),
    ("special_test", "Hawkins-Kennedy test",
     "Hawkins-Kennedy impingement test",
     "어깨", "견봉하공간/극상근", "+/-",
     "견봉하 충돌증후군",
     "충돌증후군 선별 — 민감도 74%; Neer와 조합"),
    ("special_test", "O'Brien test",
     "O'Brien active compression test",
     "어깨", "관절와순/SLAP/AC관절", "+/-",
     "SLAP 병변, AC joint 병변",
     "SLAP·AC joint 감별 — 두 검사 조합 민감도 99%"),
    ("special_test", "Elbow flexion test",
     "Elbow flexion compression test",
     "팔꿈치", "척골신경/cubital tunnel", "+/-",
     "팔꿈치터널증후군(Cubital tunnel)",
     "민감도 75% — Tinel보다 높음"),
    ("special_test", "Stork test",
     "One-leg hyperextension test (Stork)",
     "요추", "척추분리협부", "+/-",
     "척추분리증, 척추전방전위증",
     "민감도 79%/특이도 97% — 청소년 운동선수 요통 필수"),
    ("special_test", "Pace test",
     "Pace abduction test (piriformis)",
     "고관절", "이상근", "+/-",
     "이상근증후군",
     "rel_diagnosis_exam에 있으나 item 누락 — 저항 외전/외회전"),
    ("special_test", "Gaenslen test",
     "Gaenslen's SI joint provocation test",
     "척추", "천장관절", "+/-",
     "천장관절 기능장애",
     "SI joint 3개 이상 양성 시 확진율 60%↑"),
]

# ══════════════════════════════════════════════════════════════════
# 7. rel_diagnosis_exam  — 수정 1건 + 신규 추가
# ══════════════════════════════════════════════════════════════════
# (diagnosis, key_exam, positive_소견, 필요이유)
RE_R = "CCF 차팅 누락 알림 강화 — 다빈도 진단 검사 매핑"
new_rel = [
    # 어깨
    ("어깨충돌증후군",          "Neer sign, Hawkins-Kennedy test",        "양성(팔 거상 시 통증 재현)",         "충돌증후군 핵심 선별검사 2종 — 누락 시 O섹션 불완전"),
    ("SLAP 병변",               "O'Brien test",                           "1단계(내회전) 양성, 2단계(외회전) 음성", "SLAP 특이적 양성 패턴 — AC joint와 감별"),
    ("어깨 불안정성(전방)",     "Apprehension test, Relocation test",     "앞으로 빠질 것 같은 불안감 → 재위치 시 완화", RE_R),
    # 팔꿈치
    ("팔꿈치터널증후군",        "Elbow flexion test, Tinel(elbow)",       "4-5지 저림/방사통 재현",              "Elbow flexion test 민감도 75% — Tinel(54%)보다 우수"),
    ("내측상과염",              "내측상과 압통, 저항 손목 굴곡 통증",     "내측상과 국소 압통 재현",             "Golfer's elbow — 외측상과염과 대칭 추가"),
    # 척추
    ("척추분리증/전방전위증",   "Stork test(one-leg hyperextension)",     "양성(동측 요통 재현)",                "청소년 운동선수 요통 — 특이도 97%"),
    ("천장관절 기능장애",       "Gaenslen test, ASIS Compression, Thigh thrust", "3개 이상 양성",           "SI joint 복합 유발검사 — 3개↑ 양성 확진율 60%"),
    # 고관절
    ("대퇴비구충돌증후군(FAI)", "FADIR, Scour test",                      "서혜부 통증 재현",                   "고관절OA와 별도 — 젊은 환자 고관절 통증"),
    # 발목 — 아킬레스건 분리
    ("아킬레스건염",            "아킬레스건 압통 촉진, Royal London Hospital test", "건 압통·부하시 통증 재현", "건염 진단 — Thompson은 완전파열용이므로 별도 매핑"),
    # 기타
    ("중둔근 약화/GTPS",        "Trendelenburg test",                     "양성(골반 반대측 하강)",              "보행 중 골반 불안정 — 중둔근 TPI 전 필수 확인"),
]

# ══════════════════════════════════════════════════════════════════
# 8. body_part  — 새 노드 ID 4001~
# ══════════════════════════════════════════════════════════════════
# (id, parent_id, name_ko, name_en, node_type, region, midline, 필요이유, description)
BP_R_KNEE   = "무릎 주변 임상 다빈도 압통부위 — 기존 노드만으로 특정 불가"
BP_R_SHLD   = "어깨 낭·건 부착부 — 충돌증후군·석회화 압통점 구체화"
BP_R_ANKLE  = "발목 주변 건·낭 — 비골건/후경골건 건염 위치 특정"
BP_R_HIP    = "고관절 주변 낭·근 — 대전자낭·중둔근 TPI 위치"
BP_R_CERV   = "경추 주변 근육 — 흉곽출구증후군·경추성두통 압통점"
new_body = [
    # 무릎 추가 (parent=6)
    (4001, 6,    "장경인대(IT band)/외측대퇴상과",  "IT band / Lateral femoral epicondyle",  "구조물", "무릎",  False, BP_R_KNEE, "IT band / Lateral femoral epicondyle (구조물)"),
    (4002, 6,    "슬개하지방패드(Hoffa's fat pad)",  "Infrapatellar (Hoffa's) fat pad",       "구조물", "무릎",  False, BP_R_KNEE, "Hoffa's fat pad (구조물)"),
    (4003, 6,    "슬개상낭(suprapatellar pouch)",    "Suprapatellar pouch",                   "구조물", "무릎",  False, "삼출액 촉진 — Ballottement 참조점", "Suprapatellar pouch (구조물)"),
    (4004, 6,    "오금(popliteal fossa)",            "Popliteal fossa",                       "영역",   "무릎",  False, "Baker's cyst·PCL 후방 압통", "Popliteal fossa (영역)"),
    (4005, 621,  "비복근-가자미근 접합부",           "Gastrocnemius-soleus junction",         "구조물", "무릎",  False, "종아리 근육손상 위치 특정", "Gastrocnemius-soleus junction (구조물)"),
    # 어깨 추가 (parent=2)
    (4010, 2,    "견봉하낭(subacromial bursa)",      "Subacromial bursa",                     "구조물", "어깨",  False, BP_R_SHLD, "Subacromial bursa (구조물)"),
    (4011, 2,    "삼각근하낭(subdeltoid bursa)",     "Subdeltoid bursa",                      "구조물", "어깨",  False, BP_R_SHLD, "Subdeltoid bursa (구조물)"),
    (4012, 2,    "소결절(lesser tuberosity)",        "Lesser tuberosity",                     "구조물", "어깨",  False, "견갑하근건 부착점 — 내회전근 파열 압통", "Lesser tuberosity (구조물)"),
    (4013, 2,    "후방 관절낭",                      "Posterior glenohumeral capsule",        "구조물", "어깨",  False, "후방 어깨 압통·내회전 제한 OA", "Posterior glenohumeral capsule (구조물)"),
    # 발목/발 추가 (parent=7)
    (4020, 7,    "비골건(peroneal tendon sheath)",   "Peroneal tendon sheath",                "구조물", "족부",  False, BP_R_ANKLE, "Peroneal tendon sheath (구조물)"),
    (4021, 7,    "후경골건(posterior tibialis tendon)","Posterior tibialis tendon",           "구조물", "족부",  False, BP_R_ANKLE, "Posterior tibialis tendon (구조물)"),
    (4022, 7,    "전경골건(anterior tibialis tendon)","Anterior tibialis tendon",             "구조물", "족부",  False, "발목 전방 건염", "Anterior tibialis tendon (구조물)"),
    (4023, 7,    "중족지관절 2-4번(MTP 2-4)",        "2nd-4th MTP joint",                     "레벨",   "족부",  False, "Morton neuroma·metatarsalgia 압통", "2nd-4th MTP joint (레벨)"),
    (4024, 7,    "종골 후방/재트로칸테르 낭",        "Posterior calcaneus / retrocalcaneal bursa","구조물","족부",False, "아킬레스건 부착부 직전 낭 — Haglund deformity", "Retrocalcaneal bursa (구조물)"),
    # 고관절 추가 (parent=5)
    (4030, 5,    "전자낭(trochanteric bursa)",       "Trochanteric bursa",                    "구조물", "고관절",False, BP_R_HIP, "Trochanteric bursa (구조물)"),
    (4031, 5,    "장요근건/장요근낭",                "Iliopsoas tendon / Iliopsoas bursa",    "구조물", "고관절",False, "서혜부 심부 통증 — snapping hip", "Iliopsoas tendon-bursa (구조물)"),
    # 근육 MPS 추가 (parent=1000)
    (4040, 1000, "중둔근(gluteus medius)",           "Gluteus medius",                        "근육",   "근육",  False, "이상근과 별도 — 중둔근 TPI 다빈도 대상", "Gluteus medius (근육)"),
    (4041, 1000, "소둔근(gluteus minimus)",          "Gluteus minimus",                       "근육",   "근육",  False, "대전자 주위 통증 — 중둔근 하방", "Gluteus minimus (근육)"),
    # 경추 주변 근육 (parent=10)
    (4050, 10,   "전사각근(anterior scalene)",       "Anterior scalene",                      "근육",   "척추",  False, BP_R_CERV, "Anterior scalene (근육)"),
    (4051, 10,   "중사각근(middle scalene)",         "Middle scalene",                        "근육",   "척추",  False, BP_R_CERV, "Middle scalene (근육)"),
    (4052, 10,   "두반극근(semispinalis capitis)",   "Semispinalis capitis",                  "근육",   "척추",  False, "경추성 두통 기원 근육 — 후두하부 압통", "Semispinalis capitis (근육)"),
]


# ══════════════════════════════════════════════════════════════════
# 실행
# ══════════════════════════════════════════════════════════════════

def append_rows(wb, sheet_name, new_rows):
    ws = wb[sheet_name]
    before = ws.max_row - 1
    for row in new_rows:
        ws.append(list(row))
    after = ws.max_row - 1
    print(f"  {sheet_name}: {before}행 → {after}행 (+{after-before}행)")

def fix_cell(wb, sheet_name, search_col_idx, search_val, target_col_idx, new_val):
    """특정 셀 값을 찾아 다른 컬럼값을 수정."""
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2):
        if str(row[search_col_idx].value or "").strip() == search_val.strip():
            old = row[target_col_idx].value
            row[target_col_idx].value = new_val
            print(f"  [{sheet_name}] 수정: '{old}' → '{new_val}'")
            return
    print(f"  [{sheet_name}] 경고: '{search_val}' 행을 찾지 못함")

wb = load_workbook(SRC)

print("[보완 필요 시트 추가 시작]")
append_rows(wb, "rom_reference",     new_rom)
append_rows(wb, "grading_scale",     new_grading)
append_rows(wb, "outcome_scale",     new_outcome)
append_rows(wb, "peripheral_nerve",  new_nerve)
append_rows(wb, "charting_abbrev",   new_abbrev)
append_rows(wb, "exam_item",         new_exam)
append_rows(wb, "body_part",         new_body)

# rel_diagnosis_exam: 아킬레스건염→Thompson 오매핑 수정
print("\n[rel_diagnosis_exam 오매핑 수정]")
# diagnosis 컬럼(0), key_exam 컬럼(1)
fix_cell(wb, "rel_diagnosis_exam",
         search_col_idx=0, search_val="아킬레스건염",
         target_col_idx=0, new_val="아킬레스건 파열")  # 진단명 수정
fix_cell(wb, "rel_diagnosis_exam",
         search_col_idx=0, search_val="아킬레스건 파열",
         target_col_idx=2, new_val="종아리 짜기 시 발목 자발 저굴 없음(양성)")  # positive 소견 수정

print("\n[rel_diagnosis_exam 신규 매핑 추가]")
append_rows(wb, "rel_diagnosis_exam", new_rel)

wb.save(SRC)
print(f"\n[저장 완료] {SRC}")
